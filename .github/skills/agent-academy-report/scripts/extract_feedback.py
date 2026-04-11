#!/usr/bin/env python3
"""Extract feedback data from Excel files in the workspace data/ folder.

Usage: python3 extract_feedback.py <workspace-path>

Reads all .xlsx files from <workspace>/data/, extracts feedback text,
completion dates, and respondent names. Outputs <workspace>/data/_extracted.json.
"""

import openpyxl
import os
import sys
import re
import json
import glob
from datetime import datetime


COURSE_NAME_MAP = {
    "Agent Academy - Special Ops_ YAML Specialist": "Special Ops: YAML Specialist",
    "Agent Academy - Special Ops_Microsoft Copilot Studio ❤️ MCP Mission Completion": "Special Ops: Microsoft Copilot Studio MCP",
    "Agent Academy - Special Ops_ Microsoft Learn Docs MCP Server": "Special Ops: Microsoft Learn Docs MCP Server",
    "Agent Academy - Special Ops_⚡ Power Platform CLI MCP Server": "Special Ops: Power Platform CLI MCP Server",
    "Copilot Studio Agent Academy - Operative": "Operative",
    "Copilot Studio Agent Academy - Recruit": "Recruit",
    "Cowork Collective_ Badge Check Mission Completion": "Cowork Collective: Badge Check",
    "Cowork Collective_ Compliance Packet Mission Completion": "Cowork Collective: Compliance Packet",
    "Cowork Collective_ Out of Office Mission Completion": "Cowork Collective: Out of Office",
}

SKIP_PATTERNS = [r'^(no|none|n/?a|nothing|nope|-|\.+|\*+|ok|okay)$']


def get_course_name(filename):
    base = os.path.basename(filename)
    base = re.sub(r'\(\d+-\d+\)\.xlsx$', '', base).strip().replace('\xa0', ' ')
    return COURSE_NAME_MAP.get(base, base)


def get_name_or_email(row_dict):
    name = row_dict.get('Name')
    if name and isinstance(name, str) and name.strip() and name.strip().lower() != 'anonymous':
        return name.strip()
    for key in ['Email address', 'What is your email address?']:
        email = row_dict.get(key)
        if email and isinstance(email, str) and email.strip():
            return re.sub(r'[._]', ' ', email.strip().split('@')[0]).title()
    return "Anonymous"


def extract_file(filepath):
    """Extract records from a single Excel file."""
    course = get_course_name(filepath)
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    # Find feedback and completion time columns
    feedback_cols = [i for i, h in enumerate(headers) if h and 'feedback' in str(h).lower()]
    completion_col = next(
        (i for i, h in enumerate(headers) if h and 'completion time' in str(h).lower()),
        None
    )

    records = []

    if not feedback_cols:
        # No feedback column (Operative/Recruit) — extract completions only
        for row_num in range(2, ws.max_row + 1):
            dt = ws.cell(row=row_num, column=(completion_col or 1) + 1).value if completion_col is not None else None
            if isinstance(dt, datetime):
                records.append({
                    'course': course,
                    'date': dt.strftime('%Y-%m-%d'),
                    'month': dt.strftime('%Y-%m'),
                    'sentiment': None,
                    'text': None,
                    'name': None,
                    'type': 'completion',
                    'source': 'excel',
                })
        wb.close()
        return records

    for row_num in range(2, ws.max_row + 1):
        row_values = {h: ws.cell(row=row_num, column=ci + 1).value for ci, h in enumerate(headers) if h}
        dt = ws.cell(row=row_num, column=(completion_col or 1) + 1).value if completion_col is not None else None
        date_str = dt.strftime('%Y-%m-%d') if isinstance(dt, datetime) else None
        month_str = dt.strftime('%Y-%m') if isinstance(dt, datetime) else None
        name = get_name_or_email(row_values)

        for fc in feedback_cols:
            text = ws.cell(row=row_num, column=fc + 1).value
            if not text or not isinstance(text, str) or len(text.strip()) < 5:
                continue
            if any(re.match(sp, text.strip(), re.IGNORECASE) for sp in SKIP_PATTERNS):
                continue
            records.append({
                'course': course,
                'date': date_str,
                'month': month_str,
                'sentiment': None,
                'text': text.strip(),
                'name': name,
                'type': 'feedback',
                'source': 'excel',
            })

    wb.close()
    return records


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 extract_feedback.py <workspace-path>")
        sys.exit(1)

    workspace = sys.argv[1]
    data_dir = os.path.join(workspace, "data")

    files = sorted(glob.glob(os.path.join(data_dir, "*.xlsx")))
    if not files:
        print(f"No .xlsx files found in {data_dir}")
        sys.exit(1)

    all_records = []
    for f in files:
        course = get_course_name(f)
        records = extract_file(f)
        all_records.extend(records)
        feedback_count = sum(1 for r in records if r['type'] == 'feedback')
        completion_count = sum(1 for r in records if r['type'] == 'completion')
        print(f"  {course}: {feedback_count} feedback, {completion_count} completions")

    output_path = os.path.join(data_dir, "_extracted.json")
    with open(output_path, 'w') as f:
        json.dump(all_records, f, indent=2, default=str)

    total_feedback = sum(1 for r in all_records if r['type'] == 'feedback')
    total_completions = sum(1 for r in all_records if r['type'] == 'completion')
    print(f"\nTotal: {total_feedback} feedback items, {total_completions} completions")
    print(f"Saved to {output_path}")


if __name__ == "__main__":
    main()
